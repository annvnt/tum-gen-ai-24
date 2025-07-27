"""
Comprehensive trial balance processing service with validation and Excel import.
Handles trial balance validation, error detection, and processing for financial statements.
"""

import pandas as pd
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import List, Dict, Optional, Tuple, Any
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from financial_analysis.models.accounting_models import (
    TrialBalance, TrialBalanceAccount, AccountType, AccountSubType
)

logger = logging.getLogger(__name__)


class TrialBalanceValidationError(Exception):
    """Custom exception for trial balance validation errors."""
    
    def __init__(self, message: str, error_type: str, details: Dict[str, Any] = None):
        super().__init__(message)
        self.error_type = error_type
        self.details = details or {}


class TrialBalanceProcessor:
    """Service for processing and validating trial balances."""
    
    # Standard account mapping for common account codes
    ACCOUNT_MAPPING = {
        # Assets - Current
        '1000': (AccountType.ASSET, AccountSubType.CURRENT_ASSET, "Cash"),
        '1010': (AccountType.ASSET, AccountSubType.CURRENT_ASSET, "Cash - Operating"),
        '1020': (AccountType.ASSET, AccountSubType.CURRENT_ASSET, "Cash - Payroll"),
        '1100': (AccountType.ASSET, AccountSubType.CURRENT_ASSET, "Accounts Receivable"),
        '1110': (AccountType.ASSET, AccountSubType.CURRENT_ASSET, "Allowance for Doubtful Accounts"),
        '1200': (AccountType.ASSET, AccountSubType.CURRENT_ASSET, "Inventory"),
        '1210': (AccountType.ASSET, AccountSubType.CURRENT_ASSET, "Raw Materials Inventory"),
        '1220': (AccountType.ASSET, AccountSubType.CURRENT_ASSET, "Work in Process Inventory"),
        '1230': (AccountType.ASSET, AccountSubType.CURRENT_ASSET, "Finished Goods Inventory"),
        '1300': (AccountType.ASSET, AccountSubType.CURRENT_ASSET, "Prepaid Expenses"),
        '1310': (AccountType.ASSET, AccountSubType.CURRENT_ASSET, "Prepaid Insurance"),
        '1320': (AccountType.ASSET, AccountSubType.CURRENT_ASSET, "Prepaid Rent"),
        
        # Assets - Non-Current
        '1500': (AccountType.ASSET, AccountSubType.PROPERTY_PLANT_EQUIPMENT, "Land"),
        '1510': (AccountType.ASSET, AccountSubType.PROPERTY_PLANT_EQUIPMENT, "Buildings"),
        '1520': (AccountType.ASSET, AccountSubType.PROPERTY_PLANT_EQUIPMENT, "Equipment"),
        '1530': (AccountType.ASSET, AccountSubType.PROPERTY_PLANT_EQUIPMENT, "Vehicles"),
        '1540': (AccountType.ASSET, AccountSubType.PROPERTY_PLANT_EQUIPMENT, "Accumulated Depreciation - Buildings"),
        '1550': (AccountType.ASSET, AccountSubType.PROPERTY_PLANT_EQUIPMENT, "Accumulated Depreciation - Equipment"),
        '1560': (AccountType.ASSET, AccountSubType.PROPERTY_PLANT_EQUIPMENT, "Accumulated Depreciation - Vehicles"),
        '1600': (AccountType.ASSET, AccountSubType.INTANGIBLE_ASSET, "Patents"),
        '1610': (AccountType.ASSET, AccountSubType.INTANGIBLE_ASSET, "Copyrights"),
        '1620': (AccountType.ASSET, AccountSubType.INTANGIBLE_ASSET, "Trademarks"),
        '1630': (AccountType.ASSET, AccountSubType.INTANGIBLE_ASSET, "Goodwill"),
        '1700': (AccountType.ASSET, AccountSubType.INVESTMENT, "Long-term Investments"),
        
        # Liabilities - Current
        '2000': (AccountType.LIABILITY, AccountSubType.CURRENT_LIABILITY, "Accounts Payable"),
        '2100': (AccountType.LIABILITY, AccountSubType.CURRENT_LIABILITY, "Accrued Expenses"),
        '2110': (AccountType.LIABILITY, AccountSubType.CURRENT_LIABILITY, "Accrued Salaries"),
        '2120': (AccountType.LIABILITY, AccountSubType.CURRENT_LIABILITY, "Accrued Interest"),
        '2200': (AccountType.LIABILITY, AccountSubType.CURRENT_LIABILITY, "Short-term Notes Payable"),
        '2300': (AccountType.LIABILITY, AccountSubType.CURRENT_LIABILITY, "Current Portion of Long-term Debt"),
        '2400': (AccountType.LIABILITY, AccountSubType.CURRENT_LIABILITY, "Income Tax Payable"),
        
        # Liabilities - Non-Current
        '2500': (AccountType.LIABILITY, AccountSubType.NON_CURRENT_LIABILITY, "Long-term Notes Payable"),
        '2510': (AccountType.LIABILITY, AccountSubType.NON_CURRENT_LIABILITY, "Mortgage Payable"),
        '2520': (AccountType.LIABILITY, AccountSubType.NON_CURRENT_LIABILITY, "Bonds Payable"),
        '2530': (AccountType.LIABILITY, AccountSubType.NON_CURRENT_LIABILITY, "Deferred Tax Liability"),
        
        # Equity
        '3000': (AccountType.EQUITY, AccountSubType.PAID_IN_CAPITAL, "Common Stock"),
        '3010': (AccountType.EQUITY, AccountSubType.PAID_IN_CAPITAL, "Preferred Stock"),
        '3020': (AccountType.EQUITY, AccountSubType.PAID_IN_CAPITAL, "Additional Paid-in Capital"),
        '3100': (AccountType.EQUITY, AccountSubType.RETAINED_EARNINGS, "Retained Earnings"),
        '3200': (AccountType.EQUITY, AccountSubType.TREASURY_STOCK, "Treasury Stock"),
        
        # Revenue
        '4000': (AccountType.REVENUE, AccountSubType.OPERATING_REVENUE, "Sales Revenue"),
        '4010': (AccountType.REVENUE, AccountSubType.OPERATING_REVENUE, "Service Revenue"),
        '4020': (AccountType.REVENUE, AccountSubType.OPERATING_REVENUE, "Rental Revenue"),
        '4100': (AccountType.REVENUE, AccountSubType.NON_OPERATING_REVENUE, "Interest Revenue"),
        '4110': (AccountType.REVENUE, AccountSubType.NON_OPERATING_REVENUE, "Dividend Revenue"),
        '4120': (AccountType.REVENUE, AccountSubType.NON_OPERATING_REVENUE, "Gain on Sale of Assets"),
        
        # Cost of Goods Sold
        '5000': (AccountType.EXPENSE, AccountSubType.COST_OF_GOODS_SOLD, "Cost of Goods Sold"),
        '5010': (AccountType.EXPENSE, AccountSubType.COST_OF_GOODS_SOLD, "Raw Materials Used"),
        '5020': (AccountType.EXPENSE, AccountSubType.COST_OF_GOODS_SOLD, "Direct Labor"),
        '5030': (AccountType.EXPENSE, AccountSubType.COST_OF_GOODS_SOLD, "Manufacturing Overhead"),
        
        # Operating Expenses
        '6000': (AccountType.EXPENSE, AccountSubType.SELLING_EXPENSE, "Advertising Expense"),
        '6010': (AccountType.EXPENSE, AccountSubType.SELLING_EXPENSE, "Sales Commissions"),
        '6020': (AccountType.EXPENSE, AccountSubType.SELLING_EXPENSE, "Delivery Expense"),
        '6100': (AccountType.EXPENSE, AccountSubType.ADMINISTRATIVE_EXPENSE, "Salaries Expense"),
        '6110': (AccountType.EXPENSE, AccountSubType.ADMINISTRATIVE_EXPENSE, "Rent Expense"),
        '6120': (AccountType.EXPENSE, AccountSubType.ADMINISTRATIVE_EXPENSE, "Utilities Expense"),
        '6130': (AccountType.EXPENSE, AccountSubType.ADMINISTRATIVE_EXPENSE, "Insurance Expense"),
        '6140': (AccountType.EXPENSE, AccountSubType.ADMINISTRATIVE_EXPENSE, "Office Supplies Expense"),
        '6200': (AccountType.EXPENSE, AccountSubType.DEPRECIATION_EXPENSE, "Depreciation Expense"),
        '6300': (AccountType.EXPENSE, AccountSubType.INTEREST_EXPENSE, "Interest Expense"),
        '6400': (AccountType.EXPENSE, AccountSubType.TAX_EXPENSE, "Income Tax Expense"),
    }
    
    def __init__(self):
        self.validation_errors = []
    
    def process_excel_trial_balance(self, file_path: str, entity_name: str, 
                                  period_start: date, period_end: date,
                                  sheet_name: Optional[str] = None) -> TrialBalance:
        """
        Process trial balance from Excel file with comprehensive validation.
        
        Args:
            file_path: Path to Excel file
            entity_name: Name of the business entity
            period_start: Start date of accounting period
            period_end: End date of accounting period
            sheet_name: Name of sheet to process (defaults to first sheet)
            
        Returns:
            Validated TrialBalance object
            
        Raises:
            TrialBalanceValidationError: If validation fails
        """
        try:
            # Load Excel file
            df = self._load_excel_data(file_path, sheet_name)
            
            # Clean and validate data
            df = self._clean_dataframe(df)
            
            # Create trial balance accounts
            accounts = self._create_accounts_from_dataframe(df)
            
            # Create trial balance object
            trial_balance = TrialBalance(
                entity_name=entity_name,
                period_start=period_start,
                period_end=period_end,
                accounts=accounts
            )
            
            # Validate trial balance
            self._validate_trial_balance(trial_balance)
            
            if self.validation_errors:
                raise TrialBalanceValidationError(
                    "Trial balance validation failed",
                    "VALIDATION_ERROR",
                    {"errors": self.validation_errors}
                )
            
            logger.info(f"Successfully processed trial balance for {entity_name}")
            return trial_balance
            
        except Exception as e:
            logger.error(f"Error processing trial balance: {str(e)}")
            raise
    
    def _load_excel_data(self, file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """Load data from Excel file with flexible column mapping."""
        try:
            # Determine sheet to read
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                # Load first sheet
                xl_file = pd.ExcelFile(file_path)
                df = pd.read_excel(file_path, sheet_name=xl_file.sheet_names[0])
            
            return df
            
        except Exception as e:
            raise TrialBalanceValidationError(
                f"Failed to load Excel file: {str(e)}",
                "FILE_READ_ERROR"
            )
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and standardize dataframe format."""
        # Standardize column names
        column_mapping = {
            'account_code': ['account_code', 'account no', 'account number', 'code'],
            'account_name': ['account_name', 'description', 'account description', 'name'],
            'debit_balance': ['debit', 'debit_balance', 'debits', 'debit amount'],
            'credit_balance': ['credit', 'credit_balance', 'credits', 'credit amount'],
            'account_type': ['account_type', 'type', 'category'],
        }
        
        # Find actual column names
        actual_columns = {}
        df.columns = df.columns.str.strip().str.lower()
        
        for standard_name, possible_names in column_mapping.items():
            for col in df.columns:
                if col in [name.lower() for name in possible_names]:
                    actual_columns[standard_name] = col
                    break
        
        # Rename columns to standard names
        df = df.rename(columns={v: k for k, v in actual_columns.items()})
        
        # Ensure required columns exist
        required_columns = ['account_code', 'account_name']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise TrialBalanceValidationError(
                f"Missing required columns: {missing_columns}",
                "MISSING_COLUMNS",
                {"missing_columns": missing_columns, "available_columns": df.columns.tolist()}
            )
        
        # Handle missing optional columns
        if 'debit_balance' not in df.columns:
            df['debit_balance'] = 0
        if 'credit_balance' not in df.columns:
            df['credit_balance'] = 0
        
        # Clean data
        df['account_code'] = df['account_code'].astype(str).str.strip()
        df['account_name'] = df['account_name'].astype(str).str.strip()
        
        # Convert numeric columns
        numeric_columns = ['debit_balance', 'credit_balance']
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
        # Remove empty rows
        df = df[df['account_code'] != '']
        df = df[df['account_name'] != '']
        
        return df
    
    def _create_accounts_from_dataframe(self, df: pd.DataFrame) -> List[TrialBalanceAccount]:
        """Create trial balance accounts from dataframe."""
        accounts = []
        
        for _, row in df.iterrows():
            account_code = str(row['account_code']).strip()
            account_name = str(row['account_name']).strip()
            
            # Determine account type and subtype
            account_type, account_subtype = self._determine_account_type(account_code, account_name)
            
            # Get balances
            debit_balance = Decimal(str(row.get('debit_balance', 0))).quantize(
                Decimal('0.01'), rounding=ROUND_HALF_UP
            )
            credit_balance = Decimal(str(row.get('credit_balance', 0))).quantize(
                Decimal('0.01'), rounding=ROUND_HALF_UP
            )
            
            # Skip zero-balance accounts
            if debit_balance == 0 and credit_balance == 0:
                continue
            
            account = TrialBalanceAccount(
                account_code=account_code,
                account_name=account_name,
                account_type=account_type,
                account_subtype=account_subtype,
                debit_balance=debit_balance if debit_balance > 0 else None,
                credit_balance=credit_balance if credit_balance > 0 else None
            )
            
            accounts.append(account)
        
        return accounts
    
    def _determine_account_type(self, account_code: str, account_name: str) -> Tuple[AccountType, AccountSubType]:
        """Determine account type and subtype based on code or name."""
        # Check standard mapping first
        if account_code in self.ACCOUNT_MAPPING:
            account_type, account_subtype, _ = self.ACCOUNT_MAPPING[account_code]
            return account_type, account_subtype
        
        # Determine from account code number ranges
        try:
            code_num = int(account_code)
            
            # Asset accounts (1000-1999)
            if 1000 <= code_num < 2000:
                if code_num < 1500:
                    return AccountType.ASSET, AccountSubType.CURRENT_ASSET
                elif code_num < 1600:
                    return AccountType.ASSET, AccountSubType.PROPERTY_PLANT_EQUIPMENT
                elif code_num < 1700:
                    return AccountType.ASSET, AccountSubType.INTANGIBLE_ASSET
                else:
                    return AccountType.ASSET, AccountSubType.NON_CURRENT_ASSET
            
            # Liability accounts (2000-2999)
            elif 2000 <= code_num < 3000:
                if code_num < 2500:
                    return AccountType.LIABILITY, AccountSubType.CURRENT_LIABILITY
                else:
                    return AccountType.LIABILITY, AccountSubType.NON_CURRENT_LIABILITY
            
            # Equity accounts (3000-3999)
            elif 3000 <= code_num < 4000:
                if 'treasury' in account_name.lower():
                    return AccountType.EQUITY, AccountSubType.TREASURY_STOCK
                elif 'retained' in account_name.lower():
                    return AccountType.EQUITY, AccountSubType.RETAINED_EARNINGS
                else:
                    return AccountType.EQUITY, AccountSubType.PAID_IN_CAPITAL
            
            # Revenue accounts (4000-4999)
            elif 4000 <= code_num < 5000:
                return AccountType.REVENUE, AccountSubType.OPERATING_REVENUE
            
            # Cost of goods sold (5000-5999)
            elif 5000 <= code_num < 6000:
                return AccountType.EXPENSE, AccountSubType.COST_OF_GOODS_SOLD
            
            # Operating expenses (6000-6999)
            elif 6000 <= code_num < 7000:
                account_name_lower = account_name.lower()
                if 'depreciation' in account_name_lower:
                    return AccountType.EXPENSE, AccountSubType.DEPRECIATION_EXPENSE
                elif 'interest' in account_name_lower:
                    return AccountType.EXPENSE, AccountSubType.INTEREST_EXPENSE
                elif 'tax' in account_name_lower:
                    return AccountType.EXPENSE, AccountSubType.TAX_EXPENSE
                elif 'selling' in account_name_lower or 'sales' in account_name_lower:
                    return AccountType.EXPENSE, AccountSubType.SELLING_EXPENSE
                elif 'admin' in account_name_lower:
                    return AccountType.EXPENSE, AccountSubType.ADMINISTRATIVE_EXPENSE
                else:
                    return AccountType.EXPENSE, AccountSubType.OPERATING_EXPENSE
            
        except (ValueError, TypeError):
            pass
        
        # Determine from account name keywords
        account_name_lower = account_name.lower()
        
        # Assets
        if any(keyword in account_name_lower for keyword in ['cash', 'bank', 'petty']):
            return AccountType.ASSET, AccountSubType.CURRENT_ASSET
        elif any(keyword in account_name_lower for keyword in ['receivable', 'ar']):
            return AccountType.ASSET, AccountSubType.CURRENT_ASSET
        elif any(keyword in account_name_lower for keyword in ['inventory', 'stock']):
            return AccountType.ASSET, AccountSubType.CURRENT_ASSET
        elif any(keyword in account_name_lower for keyword in ['prepaid', 'advance']):
            return AccountType.ASSET, AccountSubType.CURRENT_ASSET
        elif any(keyword in account_name_lower for keyword in ['land', 'building', 'equipment', 'vehicle']):
            return AccountType.ASSET, AccountSubType.PROPERTY_PLANT_EQUIPMENT
        elif any(keyword in account_name_lower for keyword in ['patent', 'copyright', 'trademark', 'goodwill']):
            return AccountType.ASSET, AccountSubType.INTANGIBLE_ASSET
        
        # Liabilities
        elif any(keyword in account_name_lower for keyword in ['payable', 'ap']):
            return AccountType.LIABILITY, AccountSubType.CURRENT_LIABILITY
        elif any(keyword in account_name_lower for keyword in ['accrued', 'wages', 'salaries']):
            return AccountType.LIABILITY, AccountSubType.CURRENT_LIABILITY
        elif any(keyword in account_name_lower for keyword in ['notes payable', 'loan', 'mortgage']):
            return AccountType.LIABILITY, AccountSubType.NON_CURRENT_LIABILITY
        elif any(keyword in account_name_lower for keyword in ['deferred', 'tax']):
            return AccountType.LIABILITY, AccountSubType.NON_CURRENT_LIABILITY
        
        # Equity
        elif any(keyword in account_name_lower for keyword in ['common stock', 'preferred stock', 'capital']):
            return AccountType.EQUITY, AccountSubType.PAID_IN_CAPITAL
        elif any(keyword in account_name_lower for keyword in ['retained earnings', 'retained']):
            return AccountType.EQUITY, AccountSubType.RETAINED_EARNINGS
        elif any(keyword in account_name_lower for keyword in ['treasury stock']):
            return AccountType.EQUITY, AccountSubType.TREASURY_STOCK
        
        # Revenue
        elif any(keyword in account_name_lower for keyword in ['sales', 'revenue', 'service']):
            return AccountType.REVENUE, AccountSubType.OPERATING_REVENUE
        elif any(keyword in account_name_lower for keyword in ['interest', 'dividend', 'gain']):
            return AccountType.REVENUE, AccountSubType.NON_OPERATING_REVENUE
        
        # Expenses
        elif any(keyword in account_name_lower for keyword in ['cost of goods', 'cogs']):
            return AccountType.EXPENSE, AccountSubType.COST_OF_GOODS_SOLD
        elif any(keyword in account_name_lower for keyword in ['depreciation']):
            return AccountType.EXPENSE, AccountSubType.DEPRECIATION_EXPENSE
        elif any(keyword in account_name_lower for keyword in ['interest expense']):
            return AccountType.EXPENSE, AccountSubType.INTEREST_EXPENSE
        elif any(keyword in account_name_lower for keyword in ['tax']):
            return AccountType.EXPENSE, AccountSubType.TAX_EXPENSE
        
        # Default to operating expense
        return AccountType.EXPENSE, AccountSubType.OPERATING_EXPENSE
    
    def _validate_trial_balance(self, trial_balance: TrialBalance) -> None:
        """Comprehensive validation of trial balance."""
        self.validation_errors = []
        
        # Check if balanced
        if not trial_balance.is_balanced:
            difference = abs(trial_balance.total_debits - trial_balance.total_credits)
            self.validation_errors.append({
                "type": "UNBALANCED_TRIAL_BALANCE",
                "message": f"Trial balance is not balanced. Difference: ${difference}",
                "details": {
                    "total_debits": str(trial_balance.total_debits),
                    "total_credits": str(trial_balance.total_credits),
                    "difference": str(difference)
                }
            })
        
        # Check for duplicate account codes
        account_codes = [acc.account_code for acc in trial_balance.accounts]
        duplicates = [code for code in set(account_codes) if account_codes.count(code) > 1]
        if duplicates:
            self.validation_errors.append({
                "type": "DUPLICATE_ACCOUNT_CODES",
                "message": f"Duplicate account codes found: {duplicates}",
                "details": {"duplicates": duplicates}
            })
        
        # Check for missing critical accounts
        critical_accounts = [
            ('3000', 'Common Stock'),
            ('3100', 'Retained Earnings'),
            ('1000', 'Cash')
        ]
        
        existing_codes = {acc.account_code for acc in trial_balance.accounts}
        missing_critical = [name for code, name in critical_accounts if code not in existing_codes]
        
        if missing_critical:
            self.validation_errors.append({
                "type": "MISSING_CRITICAL_ACCOUNTS",
                "message": f"Missing critical accounts: {missing_critical}",
                "details": {"missing_accounts": missing_critical}
            })
        
        # Validate account balances
        for account in trial_balance.accounts:
            if account.debit_balance and account.credit_balance:
                self.validation_errors.append({
                    "type": "AMBIGUOUS_BALANCE",
                    "message": f"Account {account.account_code} has both debit and credit balances",
                    "details": {
                        "account_code": account.account_code,
                        "account_name": account.account_name,
                        "debit_balance": str(account.debit_balance),
                        "credit_balance": str(account.credit_balance)
                    }
                })
        
        # Check for negative balances
        for account in trial_balance.accounts:
            if account.net_balance < 0 and account.account_type not in [AccountType.LIABILITY, AccountType.EQUITY]:
                self.validation_errors.append({
                    "type": "NEGATIVE_BALANCE",
                    "message": f"Account {account.account_code} has negative balance",
                    "details": {
                        "account_code": account.account_code,
                        "account_name": account.account_name,
                        "balance": str(account.net_balance)
                    }
                })
    
    def get_validation_summary(self) -> Dict[str, Any]:
        """Get summary of validation errors."""
        return {
            "has_errors": len(self.validation_errors) > 0,
            "error_count": len(self.validation_errors),
            "errors": self.validation_errors
        }
    
    def export_to_excel(self, trial_balance: TrialBalance, output_path: str) -> None:
        """Export trial balance to formatted Excel file."""
        try:
            # Create dataframe
            data = []
            for account in trial_balance.accounts:
                data.append({
                    'Account Code': account.account_code,
                    'Account Name': account.account_name,
                    'Account Type': account.account_type.value,
                    'Account Subtype': account.account_subtype.value,
                    'Debit Balance': float(account.debit_balance or 0),
                    'Credit Balance': float(account.credit_balance or 0),
                    'Net Balance': float(account.net_balance)
                })
            
            df = pd.DataFrame(data)
            
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Main trial balance
                df.to_excel(writer, sheet_name='Trial Balance', index=False)
                
                # Summary sheet
                summary_data = [
                    ['Entity Name:', trial_balance.entity_name],
                    ['Period Start:', trial_balance.period_start.strftime('%Y-%m-%d')],
                    ['Period End:', trial_balance.period_end.strftime('%Y-%m-%d')],
                    ['', ''],
                    ['Total Debits:', float(trial_balance.total_debits)],
                    ['Total Credits:', float(trial_balance.total_credits)],
                    ['', ''],
                    ['Is Balanced:', 'Yes' if trial_balance.is_balanced else 'No'],
                    ['Difference:', float(abs(trial_balance.total_debits - trial_balance.total_credits))]
                ]
                
                summary_df = pd.DataFrame(summary_data, columns=['Description', 'Value'])
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Format the Excel file
                workbook = writer.book
                
                # Format trial balance sheet
                ws = writer.sheets['Trial Balance']
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Format summary sheet
                ws = writer.sheets['Summary']
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 20
            
            logger.info(f"Trial balance exported to {output_path}")
            
        except Exception as e:
            logger.error(f"Error exporting trial balance: {str(e)}")
            raise TrialBalanceValidationError(
                f"Failed to export trial balance: {str(e)}",
                "EXPORT_ERROR"
            )