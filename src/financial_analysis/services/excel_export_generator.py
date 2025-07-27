"""
Excel Export Generator - Creates professional Excel reports from financial analysis
Uses working analysis data directly, bypassing PDF complexity
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import tempfile
import os
from typing import Dict, Any

from ..storage.gcs_client import GCSClient
from ..storage.database_manager import db_manager

class ExcelExportGenerator:
    """Simple Excel report generator using working financial analysis data"""
    
    def __init__(self):
        self.gcs_client = GCSClient()
        
    async def generate_excel_report(self, file_id: str, analysis_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate Excel report with 4 worksheets"""
        
        file_info = db_manager.get_uploaded_file(file_id)
        if not file_info:
            raise ValueError(f"File {file_id} not found")
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Create 4 worksheets
        self._create_balance_sheet(wb, analysis_data)
        self._create_income_statement(wb, analysis_data)
        self._create_cash_flow(wb, analysis_data)
        self._create_summary(wb, analysis_data, file_info['filename'])
        
        # Save and upload
        temp_dir = tempfile.mkdtemp()
        excel_path = os.path.join(temp_dir, f"report_{file_id}.xlsx")
        wb.save(excel_path)
        
        gcs_url = await self.gcs_client.upload_file(
            open(excel_path, 'rb'), 
            f"reports/{file_id}/financial_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        return {'excel_url': gcs_url, 'file_id': file_id}
    
    def _create_balance_sheet(self, wb, data):
        ws = wb.create_sheet("Balance Sheet")
        ws['A1'] = "Balance Sheet"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Add data from analysis
        if 'tables' in data and 'balance_sheet' in data['tables']:
            for i, row in enumerate(data['tables']['balance_sheet'], 3):
                ws[f'A{i}'] = str(row)
        else:
            ws['A3'] = "Balance Sheet Data"
    
    def _create_income_statement(self, wb, data):
        ws = wb.create_sheet("Income Statement")
        ws['A1'] = "Income Statement"
        ws['A1'].font = Font(size=16, bold=True)
        
        if 'tables' in data and 'income_statement' in data['tables']:
            for i, row in enumerate(data['tables']['income_statement'], 3):
                ws[f'A{i}'] = str(row)
        else:
            ws['A3'] = "Income Statement Data"
    
    def _create_cash_flow(self, wb, data):
        ws = wb.create_sheet("Cash Flow")
        ws['A1'] = "Cash Flow Statement"
        ws['A1'].font = Font(size=16, bold=True)
        
        if 'summary' in data:
            ws['A3'] = data['summary'][:5000]  # Summary goes here
    
    def _create_summary(self, wb, data, filename):
        ws = wb.create_sheet("Executive Summary")
        ws['A1'] = "Executive Summary & Recommendations"
        ws['A1'].font = Font(size=16, bold=True)
        
        ws['A3'] = f"Report for: {filename}"
        ws['A4'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        
        insights = self._generate_insights_text(data)
        ws['A6'] = insights
        
        # Format for better readability
        ws.column_dimensions['A'].width = 80

    def _generate_insights_text(self, data: Dict[str, Any]) -> str:
        '''Generate simple accounting insights'''
        tables = data.get('tables', {})
        
        # Extract key numbers
        income_data = tables.get('income_statement', [])
        revenue_2024 = 0
        revenue_2023 = 0
        loss_2024 = 0
        loss_2023 = 0
        
        for item in income_data:
            indicator = str(item.get(' Indicator', ''))
            try:
                if 'Revenue' in indicator:
                    revenue_2024 = float(str(item.get('Current Year', 0)).replace(',', ''))
                    revenue_2023 = float(str(item.get('Previous Year  ', 0)).replace(',', ''))
                elif 'Net Loss' in indicator:
                    loss_2024 = float(str(item.get('Current Year', 0)).replace(',', ''))
                    loss_2023 = float(str(item.get('Previous Year  ', 0)).replace(',', ''))
            except:
                pass
        
        return f'''📊 BUSINESS HEALTH SUMMARY

🚨 KEY WARNINGS:
• Revenue: ${revenue_2024:,.0f} (negative = losses)
• Net loss: ${loss_2024:,.0f}
• Revenue change: ${revenue_2024 - revenue_2023:,.0f} vs last year

💡 POSITIVE SIGNS:
• Loss improvement: ${abs(loss_2024 - loss_2023):,.0f}
• Business still operating despite challenges

⚡ IMMEDIATE ACTIONS:
1. Check why revenue is negative
2. Review largest expenses  
3. Focus on sales growth
4. Consider accounting review

{data.get('summary', '')[:1000]}'''

