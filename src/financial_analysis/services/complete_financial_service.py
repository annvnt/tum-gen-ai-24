"""
Complete financial statements generation service.
Orchestrates the generation of all financial statements from trial balance data.
"""

from datetime import date, datetime
from decimal import Decimal
from typing import Optional, Dict, Any
import logging

from financial_analysis.models.accounting_models import (
    TrialBalance, CompleteFinancialStatements, BalanceSheet, IncomeStatement,
    StatementOfEquity, CashFlowStatement, FinancialRatios
)
from financial_analysis.services.trial_balance_processor import TrialBalanceProcessor
from financial_analysis.services.balance_sheet_generator import BalanceSheetGenerator
from financial_analysis.services.income_statement_generator import IncomeStatementGenerator
from financial_analysis.services.equity_statement_generator import EquityStatementGenerator
from financial_analysis.services.cash_flow_generator import CashFlowGenerator
from financial_analysis.services.financial_ratios_calculator import FinancialRatiosCalculator

logger = logging.getLogger(__name__)


class CompleteFinancialService:
    """Service for generating complete set of financial statements."""
    
    def __init__(self):
        self.trial_balance_processor = TrialBalanceProcessor()
        self.balance_sheet_generator = BalanceSheetGenerator()
        self.income_statement_generator = IncomeStatementGenerator()
        self.equity_statement_generator = EquityStatementGenerator()
        self.cash_flow_generator = CashFlowGenerator()
        self.ratios_calculator = FinancialRatiosCalculator()
    
    def generate_complete_financial_statements(self, trial_balance: TrialBalance,
                                             beginning_cash_balance: Decimal,
                                             beginning_retained_earnings: Decimal,
                                             dividends: Optional[Decimal] = None,
                                             cash_flow_method: str = "indirect") -> CompleteFinancialStatements:
        """
        Generate complete set of financial statements from trial balance.
        
        Args:
            trial_balance: Validated trial balance
            beginning_cash_balance: Cash balance at beginning of period
            beginning_retained_earnings: Retained earnings at beginning of period
            dividends: Dividends declared during period
            cash_flow_method: "direct" or "indirect" method for cash flow
            
        Returns:
            Complete set of financial statements
        """
        
        logger.info(f"Starting complete financial statements generation for {trial_balance.entity_name}")
        
        # Generate individual statements
        balance_sheet = self.balance_sheet_generator.generate_balance_sheet(trial_balance)
        income_statement = self.income_statement_generator.generate_income_statement(trial_balance)
        
        # Generate equity statement
        equity_statement = self.equity_statement_generator.generate_equity_statement(
            trial_balance=trial_balance,
            net_income=income_statement.net_income,
            dividends=dividends,
            period_start=trial_balance.period_start,
            period_end=trial_balance.period_end
        )
        
        # Generate cash flow statement
        cash_flow_statement = self.cash_flow_generator.generate_cash_flow_statement(
            trial_balance=trial_balance,
            net_income=income_statement.net_income,
            beginning_cash_balance=beginning_cash_balance,
            method=cash_flow_method,
            period_start=trial_balance.period_start,
            period_end=trial_balance.period_end
        )
        
        # Calculate financial ratios
        financial_ratios = self.ratios_calculator.calculate_financial_ratios(
            balance_sheet=balance_sheet,
            income_statement=income_statement,
            cash_flow_statement=cash_flow_statement
        )
        
        # Create complete financial statements
        complete_statements = CompleteFinancialStatements(
            entity_name=trial_balance.entity_name,
            period_start=trial_balance.period_start,
            period_end=trial_balance.period_end,
            trial_balance=trial_balance,
            balance_sheet=balance_sheet,
            income_statement=income_statement,
            statement_of_equity=equity_statement,
            cash_flow_statement=cash_flow_statement,
            financial_ratios=financial_ratios
        )
        
        logger.info(f"Complete financial statements generated successfully for {trial_balance.entity_name}")
        return complete_statements
    
    def generate_from_excel(self, file_path: str, entity_name: str,
                          period_start: date, period_end: date,
                          beginning_cash_balance: Decimal,
                          beginning_retained_earnings: Decimal,
                          dividends: Optional[Decimal] = None,
                          cash_flow_method: str = "indirect") -> CompleteFinancialStatements:
        """
        Generate complete financial statements from Excel trial balance file.
        
        Args:
            file_path: Path to Excel trial balance file
            entity_name: Name of the business entity
            period_start: Start date of accounting period
            period_end: End date of accounting period
            beginning_cash_balance: Cash balance at beginning of period
            beginning_retained_earnings: Retained earnings at beginning of period
            dividends: Dividends declared during period
            cash_flow_method: "direct" or "indirect" method for cash flow
            
        Returns:
            Complete set of financial statements
        """
        
        # Process trial balance from Excel
        trial_balance = self.trial_balance_processor.process_excel_trial_balance(
            file_path=file_path,
            entity_name=entity_name,
            period_start=period_start,
            period_end=period_end
        )
        
        return self.generate_complete_financial_statements(
            trial_balance=trial_balance,
            beginning_cash_balance=beginning_cash_balance,
            beginning_retained_earnings=beginning_retained_earnings,
            dividends=dividends,
            cash_flow_method=cash_flow_method
        )
    
    def get_comprehensive_analysis(self, complete_statements: CompleteFinancialStatements) -> Dict[str, Any]:
        """Get comprehensive financial analysis including all statements."""
        
        # Balance sheet analysis
        balance_sheet_analysis = self.balance_sheet_generator.get_balance_sheet_analysis(
            complete_statements.balance_sheet
        )
        
        # Income statement analysis
        income_statement_analysis = self.income_statement_generator.get_income_statement_analysis(
            complete_statements.income_statement
        )
        
        # Equity statement analysis
        equity_analysis = self.equity_statement_generator.get_equity_analysis(
            complete_statements.statement_of_equity
        )
        
        # Cash flow analysis
        cash_flow_analysis = self.cash_flow_generator.get_cash_flow_analysis(
            complete_statements.cash_flow_statement
        )
        
        # Financial ratios analysis
        ratios_analysis = self.ratios_calculator.get_ratio_analysis(
            complete_statements.financial_ratios
        )
        
        return {
            "balance_sheet_analysis": balance_sheet_analysis,
            "income_statement_analysis": income_statement_analysis,
            "equity_analysis": equity_analysis,
            "cash_flow_analysis": cash_flow_analysis,
            "ratios_analysis": ratios_analysis,
            "summary": {
                "entity_name": complete_statements.entity_name,
                "period": f"{complete_statements.period_start} to {complete_statements.period_end}",
                "net_income": float(complete_statements.income_statement.net_income),
                "total_assets": float(complete_statements.balance_sheet.total_assets),
                "total_equity": float(complete_statements.balance_sheet.equity.total_amount),
                "cash_position": float(complete_statements.cash_flow_statement.ending_cash_balance)
            }
        }
    
    def export_to_excel(self, complete_statements: CompleteFinancialStatements,
                       output_path: str, include_analysis: bool = True) -> str:
        """
        Export complete financial statements to Excel format.
        
        Args:
            complete_statements: Complete set of financial statements
            output_path: Path for output Excel file
            include_analysis: Whether to include ratio analysis
            
        Returns:
            Path to exported Excel file
        """
        
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets for each statement
        self._create_balance_sheet_sheet(wb, complete_statements.balance_sheet)
        self._create_income_statement_sheet(wb, complete_statements.income_statement)
        self._create_equity_statement_sheet(wb, complete_statements.statement_of_equity)
        self._create_cash_flow_sheet(wb, complete_statements.cash_flow_statement)
        self._create_ratios_sheet(wb, complete_statements.financial_ratios)
        
        if include_analysis:
            analysis = self.get_comprehensive_analysis(complete_statements)
            self._create_analysis_sheet(wb, analysis)
        
        # Save workbook
        wb.save(output_path)
        
        logger.info(f"Financial statements exported to {output_path}")
        return output_path
    
    def _create_balance_sheet_sheet(self, wb, balance_sheet: BalanceSheet) -> None:
        """Create balance sheet worksheet."""
        ws = wb.create_sheet("Balance Sheet")
        
        # Title
        ws['A1'] = "Balance Sheet"
        ws['A2'] = balance_sheet.entity_name
        ws['A3'] = f"As of {balance_sheet.statement_date}"
        
        # Assets
        ws['A5'] = "ASSETS"
        ws['A5'].font = Font(bold=True)
        
        row = 6
        for item in balance_sheet.assets.items:
            ws[f'A{row}'] = item.account_name
            ws[f'B{row}'] = float(item.amount)
            ws[f'B{row}'].number_format = '"$"#,##0.00'
            row += 1
        
        ws[f'A{row}'] = "Total Assets"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(balance_sheet.assets.total_amount)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        
        # Liabilities
        row += 2
        ws[f'A{row}'] = "LIABILITIES"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for item in balance_sheet.liabilities.items:
            ws[f'A{row}'] = item.account_name
            ws[f'B{row}'] = float(item.amount)
            ws[f'B{row}'].number_format = '"$"#,##0.00'
            row += 1
        
        ws[f'A{row}'] = "Total Liabilities"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(balance_sheet.liabilities.total_amount)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        
        # Equity
        row += 2
        ws[f'A{row}'] = "EQUITY"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for item in balance_sheet.equity.items:
            ws[f'A{row}'] = item.account_name
            ws[f'B{row}'] = float(item.amount)
            ws[f'B{row}'].number_format = '"$"#,##0.00'
            row += 1
        
        ws[f'A{row}'] = "Total Equity"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(balance_sheet.equity.total_amount)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        
        # Format columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
    
    def _create_income_statement_sheet(self, wb, income_statement: IncomeStatement) -> None:
        """Create income statement worksheet."""
        ws = wb.create_sheet("Income Statement")
        
        # Title
        ws['A1'] = "Income Statement"
        ws['A2'] = income_statement.entity_name
        ws['A3'] = f"For the period ended {income_statement.period_end}"
        
        # Revenue
        ws['A5'] = "REVENUE"
        ws['A5'].font = Font(bold=True)
        
        row = 6
        for item in income_statement.revenue.items:
            ws[f'A{row}'] = item.account_name
            ws[f'B{row}'] = float(item.amount)
            ws[f'B{row}'].number_format = '"$"#,##0.00'
            row += 1
        
        ws[f'A{row}'] = "Total Revenue"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(income_statement.revenue.total_amount)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        
        # Cost of Goods Sold
        row += 2
        ws[f'A{row}'] = "COST OF GOODS SOLD"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for item in income_statement.cost_of_goods_sold.items:
            ws[f'A{row}'] = item.account_name
            ws[f'B{row}'] = -float(item.amount)
            ws[f'B{row}'].number_format = '"$"#,##0.00'
            row += 1
        
        ws[f'A{row}'] = "Gross Profit"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(income_statement.gross_profit)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        
        # Operating Expenses
        row += 2
        ws[f'A{row}'] = "OPERATING EXPENSES"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for item in income_statement.operating_expenses.items:
            ws[f'A{row}'] = item.account_name
            ws[f'B{row}'] = -float(item.amount)
            ws[f'B{row}'].number_format = '"$"#,##0.00'
            row += 1
        
        ws[f'A{row}'] = "Operating Income"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(income_statement.operating_income)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        
        # Net Income
        row += 2
        ws[f'A{row}'] = "NET INCOME"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(income_statement.net_income)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        
        # Format columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
    
    def _create_equity_statement_sheet(self, wb, equity_statement: StatementOfEquity) -> None:
        """Create equity statement worksheet."""
        ws = wb.create_sheet("Statement of Equity")
        
        # Title
        ws['A1'] = "Statement of Changes in Equity"
        ws['A2'] = equity_statement.entity_name
        ws['A3'] = f"For the period ended {equity_statement.period_end}"
        
        # Headers
        ws['A5'] = "Description"
        ws['B5'] = "Amount"
        ws['C5'] = "Type"
        
        for col in ['A5', 'B5', 'C5']:
            ws[col].font = Font(bold=True)
        
        row = 6
        for change in equity_statement.changes:
            ws[f'A{row}'] = change.description
            ws[f'B{row}'] = float(change.amount)
            ws[f'C{row}'] = "Addition" if change.is_addition else "Deduction"
            ws[f'B{row}'].number_format = '"$"#,##0.00'
            row += 1
        
        ws[f'A{row}'] = "Ending Equity"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(equity_statement.ending_equity)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        
        # Format columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
    
    def _create_cash_flow_sheet(self, wb, cash_flow_statement: CashFlowStatement) -> None:
        """Create cash flow statement worksheet."""
        ws = wb.create_sheet("Cash Flow Statement")
        
        # Title
        ws['A1'] = "Cash Flow Statement"
        ws['A2'] = cash_flow_statement.entity_name
        ws['A3'] = f"For the period ended {cash_flow_statement.period_end}"
        
        # Operating Activities
        ws['A5'] = "OPERATING ACTIVITIES"
        ws['A5'].font = Font(bold=True)
        
        row = 6
        for item in cash_flow_statement.operating_activities:
            ws[f'A{row}'] = item.description
            ws[f'B{row}'] = float(item.amount) if item.is_inflow else -float(item.amount)
            ws[f'B{row}'].number_format = '"$"#,##0.00'
            row += 1
        
        ws[f'A{row}'] = "Net Cash from Operating Activities"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(cash_flow_statement.net_cash_operating)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        
        # Investing Activities
        row += 2
        ws[f'A{row}'] = "INVESTING ACTIVITIES"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for item in cash_flow_statement.investing_activities:
            ws[f'A{row}'] = item.description
            ws[f'B{row}'] = float(item.amount) if item.is_inflow else -float(item.amount)
            ws[f'B{row}'].number_format = '"$"#,##0.00'
            row += 1
        
        ws[f'A{row}'] = "Net Cash from Investing Activities"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(cash_flow_statement.net_cash_investing)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        
        # Financing Activities
        row += 2
        ws[f'A{row}'] = "FINANCING ACTIVITIES"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for item in cash_flow_statement.financing_activities:
            ws[f'A{row}'] = item.description
            ws[f'B{row}'] = float(item.amount) if item.is_inflow else -float(item.amount)
            ws[f'B{row}'].number_format = '"$"#,##0.00'
            row += 1
        
        ws[f'A{row}'] = "Net Cash from Financing Activities"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(cash_flow_statement.net_cash_financing)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        
        # Net Change
        row += 2
        ws[f'A{row}'] = "NET CHANGE IN CASH"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(cash_flow_statement.net_change_in_cash)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        
        # Format columns
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
    
    def _create_ratios_sheet(self, wb, ratios: FinancialRatios) -> None:
        """Create financial ratios worksheet."""
        ws = wb.create_sheet("Financial Ratios")
        
        # Title
        ws['A1'] = "Financial Ratios"
        ws['A2'] = ratios.entity_name
        ws['A3'] = f"As of {ratios.statement_date}"
        
        # Liquidity Ratios
        ws['A5'] = "LIQUIDITY RATIOS"
        ws['A5'].font = Font(bold=True)
        
        ratios_data = [
            ("Current Ratio", float(ratios.current_ratio), "x"),
            ("Quick Ratio", float(ratios.quick_ratio), "x"),
            ("Cash Ratio", float(ratios.cash_ratio), "x"),
            ("", "", ""),
            ("PROFITABILITY RATIOS", "", ""),
            ("Gross Profit Margin", float(ratios.gross_profit_margin), "%"),
            ("Operating Profit Margin", float(ratios.operating_profit_margin), "%"),
            ("Net Profit Margin", float(ratios.net_profit_margin), "%"),
            ("Return on Assets", float(ratios.return_on_assets), "%"),
            ("Return on Equity", float(ratios.return_on_equity), "%"),
            ("", "", ""),
            ("LEVERAGE RATIOS", "", ""),
            ("Debt to Equity", float(ratios.debt_to_equity), "x"),
            ("Debt to Assets", float(ratios.debt_to_assets), "%"),
            ("Equity Multiplier", float(ratios.equity_multiplier), "x"),
            ("", "", ""),
            ("EFFICIENCY RATIOS", "", ""),
            ("Asset Turnover", float(ratios.asset_turnover), "x"),
        ]
        
        if ratios.inventory_turnover:
            ratios_data.append(("Inventory Turnover", float(ratios.inventory_turnover), "x"))
        
        if ratios.receivables_turnover:
            ratios_data.append(("Receivables Turnover", float(ratios.receivables_turnover), "x"))
        
        for i, (description, value, unit) in enumerate(ratios_data, start=6):
            ws[f'A{i}'] = description
            ws[f'B{i}'] = value
            ws[f'C{i}'] = unit
        
        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
    
    def _create_analysis_sheet(self, wb, analysis: Dict[str, Any]) -> None:
        """Create analysis worksheet."""
        ws = wb.create_sheet("Financial Analysis")
        
        # Title
        ws['A1'] = "Financial Analysis Summary"
        
        # Overall Assessment
        ws['A3'] = "Overall Financial Health"
        ws['A3'].font = Font(bold=True)
        ws['A4'] = analysis["ratios_analysis"]["overall_assessment"]["overall_health"]
        ws['A5'] = analysis["ratios_analysis"]["overall_assessment"]["recommendation"]
        
        # Key Metrics
        ws['A7'] = "Key Financial Metrics"
        ws['A7'].font = Font(bold=True)
        
        metrics_data = [
            ("Net Income", analysis["summary"]["net_income"], "$"),
            ("Total Assets", analysis["summary"]["total_assets"], "$"),
            ("Total Equity", analysis["summary"]["total_equity"], "$"),
            ("Cash Position", analysis["summary"]["cash_position"], "$"),
            ("Current Ratio", analysis["balance_sheet_analysis"]["current_ratio"], "x"),
            ("Net Profit Margin", analysis["income_statement_analysis"]["net_profit_margin"], "%"),
            ("ROE", analysis["income_statement_analysis"]["return_on_equity"], "%"),
            ("Debt-to-Equity", analysis["balance_sheet_analysis"]["debt_to_equity"], "x"),
        ]
        
        for i, (metric, value, unit) in enumerate(metrics_data, start=8):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'C{i}'] = unit
        
        # Format columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 5