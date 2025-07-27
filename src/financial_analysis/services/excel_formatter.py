"""
Enhanced Excel Formatter - Creates professional Excel reports
with proper tables and formatting
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import tempfile
import os
from typing import Dict, Any

class ExcelFormatter:
    """Create professional Excel reports with proper formatting"""
    
    def create_enhanced_excel(self, analysis_data: Dict[str, Any], filename: str) -> str:
        """Create formatted Excel workbook with insights"""
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Create properly formatted worksheets
        self._create_balance_sheet_sheet(wb, analysis_data)
        self._create_income_statement_sheet(wb, analysis_data)
        self._create_cash_flow_sheet(wb, analysis_data)
        self._create_executive_summary_sheet(wb, analysis_data, filename)
        
        # Save to temporary file
        temp_dir = tempfile.mkdtemp()
        excel_path = os.path.join(temp_dir, f"enhanced_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(excel_path)
        
        return excel_path
    
    def _create_balance_sheet_sheet(self, wb, data):
        """Create formatted balance sheet"""
        ws = wb.create_sheet("Balance Sheet")
        
        # Header
        ws['A1'] = "Balance Sheet"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = ""
        
        # Process data into proper table
        bs_data = data.get('tables', {}).get('balance_sheet', [])
        if bs_data:
            # Create table data
            rows = []
            for item in bs_data:
                indicator = str(item.get(' Indicator', '')).strip()
                current = str(item.get('Current Year', '')).strip()
                previous = str(item.get('Previous Year  ', '')).strip()
                if indicator and 'Indicator' not in indicator and current != 'Current Year':
                    rows.append([indicator, current, previous])
            
            if rows:
                # Headers
                ws['A3'] = "Account"
                ws['B3'] = "Current Year"
                ws['C3'] = "Previous Year"
                
                # Style headers
                for col in ['A3', 'B3', 'C3']:
                    ws[col].font = Font(bold=True)
                    ws[col].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    ws[col].font = Font(color="FFFFFF", bold=True)
                
                # Add data
                for idx, row in enumerate(rows, 4):
                    ws[f'A{idx}'] = row[0]
                    ws[f'B{idx}'] = row[1]
                    ws[f'C{idx}'] = row[2]
        
        # Auto-fit columns
        for col in ['A', 'B', 'C']:
            ws.column_dimensions[col].width = 25
    
    def _create_income_statement_sheet(self, wb, data):
        """Create formatted income statement"""
        ws = wb.create_sheet("Income Statement")
        
        ws['A1'] = "Income Statement"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = ""
        
        income_data = data.get('tables', {}).get('income_statement', [])
        if income_data:
            rows = []
            for item in income_data:
                indicator = str(item.get(' Indicator', '')).strip()
                current = str(item.get('Current Year', '')).strip()
                previous = str(item.get('Previous Year  ', '')).strip()
                if indicator and 'Indicator' not in indicator and current != 'Current Year':
                    rows.append([indicator, current, previous])
            
            if rows:
                ws['A3'] = "Account"
                ws['B3'] = "Current Year"
                ws['C3'] = "Previous Year"
                
                for col in ['A3', 'B3', 'C3']:
                    ws[col].font = Font(bold=True)
                    ws[col].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    ws[col].font = Font(color="FFFFFF", bold=True)
                
                for idx, row in enumerate(rows, 4):
                    ws[f'A{idx}'] = row[0]
                    ws[f'B{idx}'] = row[1]
                    ws[f'C{idx}'] = row[2]
        
        for col in ['A', 'B', 'C']:
            ws.column_dimensions[col].width = 25
    
    def _create_cash_flow_sheet(self, wb, data):
        """Create formatted cash flow statement"""
        ws = wb.create_sheet("Cash Flow")
        
        ws['A1'] = "Cash Flow Statement"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = ""
        
        cf_data = data.get('tables', {}).get('cash_flow_statement', [])
        if cf_data:
            rows = []
            for item in cf_data:
                indicator = str(item.get(' Indicator', '')).strip()
                current = str(item.get('Current Year', '')).strip()
                previous = str(item.get('Previous Year  ', '')).strip()
                if indicator and 'Indicator' not in indicator and current != 'Current Year':
                    rows.append([indicator, current, previous])
            
            if rows:
                ws['A3'] = "Account"
                ws['B3'] = "Current Year"
                ws['C3'] = "Previous Year"
                
                for col in ['A3', 'B3', 'C3']:
                    ws[col].font = Font(bold=True)
                    ws[col].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    ws[col].font = Font(color="FFFFFF", bold=True)
                
                for idx, row in enumerate(rows, 4):
                    ws[f'A{idx}'] = row[0]
                    ws[f'B{idx}'] = row[1]
                    ws[f'C{idx}'] = row[2]
        
        for col in ['A', 'B', 'C']:
            ws.column_dimensions[col].width = 25
    
    def _create_executive_summary_sheet(self, wb, data, filename):
        """Create executive summary with insights"""
        ws = wb.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = "Executive Summary & Business Health Report"
        ws['A1'].font = Font(size=16, bold=True)
        
        # File info
        ws['A3'] = f"Report for: {filename}"
        ws['A3'].font = Font(size=12, bold=True)
        ws['A4'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws['A5'] = ""
        
        # Generate insights
        insights = self._generate_insights(data)
        
        # Write insights
        row = 7
        for section, content in insights.items():
            ws[f'A{row}'] = section
            ws[f'A{row}'].font = Font(size=14, bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            row += 1
            
            for key, value in content.items():
                ws[f'A{row}'] = f"• {key}:"
                ws[f'B{row}'] = str(value)
                row += 1
            row += 1
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
    
    def _generate_insights(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate structured accounting insights"""
        
        tables = data.get('tables', {})
        
        # Debug print
        print(f"Available tables: {list(tables.keys())}")
        
        # Extract key numbers from actual table names
        income_data = tables.get('income_statement', [])
        balance_data = tables.get('balance_sheet', [])
        
        print(f"Income data: {len(income_data)} items")
        print(f"Balance data: {len(balance_data)} items")
        
        # Initialize values
        revenue_2024 = 0
        revenue_2023 = 0
        loss_2024 = 0
        loss_2023 = 0
        assets_2024 = 0
        assets_2023 = 0
        
        # Parse income statement
        for item in income_data:
            indicator = str(item.get(' Indicator', '')).strip()
            try:
                if 'Revenue' in indicator:
                    current = str(item.get('Current Year', '0')).strip()
                    previous = str(item.get('Previous Year  ', '0')).strip()
                    
                    # Handle NaN and empty values
                    if current and current != 'NaN' and current != 'nan':
                        revenue_2024 = float(current.replace(',', ''))
                    if previous and previous != 'NaN' and previous != 'nan':
                        revenue_2023 = float(previous.replace(',', ''))
                        
                elif 'Net Loss' in indicator or 'Net Income' in indicator:
                    current = str(item.get('Current Year', '0')).strip()
                    previous = str(item.get('Previous Year  ', '0')).strip()
                    
                    if current and current != 'NaN' and current != 'nan':
                        loss_2024 = float(current.replace(',', ''))
                    if previous and previous != 'NaN' and previous != 'nan':
                        loss_2023 = float(previous.replace(',', ''))
                        
            except Exception as e:
                print(f"Error parsing income: {e}")
        
        # Parse balance sheet
        for item in balance_data:
            indicator = str(item.get(' Indicator', '')).strip()
            try:
                if 'Total Assets' in indicator:
                    current = str(item.get('Current Year', '0')).strip()
                    previous = str(item.get('Previous Year  ', '0')).strip()
                    
                    if current and current != 'NaN' and current != 'nan':
                        assets_2024 = float(current.replace(',', ''))
                    if previous and previous != 'NaN' and previous != 'nan':
                        assets_2023 = float(previous.replace(',', ''))
                        
            except Exception as e:
                print(f"Error parsing assets: {e}")
        
        print(f"Revenue: {revenue_2024}, {revenue_2023}")
        print(f"Loss: {loss_2024}, {loss_2023}")
        print(f"Assets: {assets_2024}, {assets_2023}")
        
        # Get actual summary text
        summary_text = data.get('summary', '')[:500]
        
        return {
            "🚨 KEY WARNINGS": {
                "Revenue (Current)": f"${revenue_2024:,.0f}",
                "Revenue (Previous)": f"${revenue_2023:,.0f}",
                "Net Result": f"${loss_2024:,.0f}",
                "Revenue Change": f"${revenue_2024 - revenue_2023:,.0f}"
            },
            "💡 INSIGHTS": {
                "Business Health": "Negative revenue indicates operational challenges",
                "Asset Trend": f"${assets_2024 - assets_2023:,.0f} change in assets",
                "Action Required": "Immediate revenue and cost analysis needed"
            },
            "⚡ IMMEDIATE ACTIONS": {
                "Priority 1": "Investigate negative revenue sources",
                "Priority 2": "Review all expense categories",
                "Priority 3": "Implement cost reduction strategies",
                "Priority 4": "Consider professional accounting review"
            },
            "📊 SUMMARY": summary_text
        }
